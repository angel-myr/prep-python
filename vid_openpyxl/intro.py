
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

ruta_ventas = "C:/data/retail_dw/sales.csv"
archivo = pd.read_csv(ruta_ventas)

# print(archivo.info())

pivote = archivo.pivot_table(
    index='item_id',
    columns='store_id',
    values=['quantity', 'sum_total'],
    aggfunc='sum',
)

pivote.to_excel('sales_resume.xlsx', startrow=4, sheet_name='ReportSales')

libro = load_workbook('sales_resume.xlsx')
hoja = libro['ReportSales']

posC_min = libro.active.min_column
posC_max = libro.active.max_column
posF_min = libro.active.min_row
posF_max = libro.active.max_row

# print({posC_min, posF_min, posC_max, posF_max})

barras = BarChart()
data = Reference(
    hoja,
    min_col=posC_min + 1,
    min_row=posF_min,
    max_col=posC_max,
    max_row=posF_max
)


data_conCat = Reference(
    hoja,
    min_col=posC_min,
    min_row=posF_min,
    max_col=posC_min,
    max_row=posF_max - 1
)

# Incluyendo data con encabezados (eje X)
barras.add_data(data, titles_from_data=True)
# Incluyendo categorías (eje Y)
barras.set_categories(data_conCat)

# Agregar gráfica desde K5
hoja.add_chart(barras, 'K5')
barras.style = 5
barras.title = 'Resumen Ventas por Tienda y Producto'

libro.save('sales_resume.xlsx')